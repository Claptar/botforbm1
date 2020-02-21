import openpyxl
import timetable
import pickle

groups = []  # список расписаний для каждого курса

# считываем расписания из экселевских файлов
# меняем их на новые в каждом семе, при замене, возмножно, нужно внести правки в функция timetable.get_timetable()
kurs_1_doc = openpyxl.load_workbook('1_kurs_vesna_2020_chernovik.xlsm')
kurs_1 = kurs_1_doc.active
groups.append(timetable.get_timetable(kurs_1))

kurs_2_doc = openpyxl.load_workbook('2_kurs_vesna_2020_chernovik.xlsm')
kurs_2 = kurs_2_doc.active
groups.append(timetable.get_timetable(kurs_2))

kurs_3_doc = openpyxl.load_workbook('3_kurs_vesna_2020_chernovik.xlsm')
kurs_3 = kurs_3_doc.active
groups.append(timetable.get_timetable(kurs_3))

kurs_4_doc = openpyxl.load_workbook('4_kurs_vesna_2020_chernovik.xlsm')
kurs_4 = kurs_4_doc.active
groups.append(timetable.get_timetable(kurs_4))

# kurs_5_doc = openpyxl.load_workbook('Raspisanie_5_kurs_osen_2019.xlsm')
# kurs_5 = kurs_5_doc.active
# groups.append(timetable.get_timetable(kurs_5))

# kurs_6_faki_doc = openpyxl.load_workbook('Raspisanie_6_kurs_FAKI_osen_2019.xlsm')
# kurs_6_faki = kurs_6_faki_doc.active
# groups_6_faki = timetable.get_timetable(kurs_6_faki)

# kurs_6_fupm_doc = openpyxl.load_workbook('Raspisanie_6_kurs_FUPM_osen_2019.xlsm')
# kurs_6_fupm = kurs_6_fupm_doc.active
# groups_6_fupm = timetable.get_timetable(kurs_6_fupm)

# groups.append({**groups_6_faki, **groups_6_fupm})

for i in range(len(groups)):  # запись расписаний в удобный для хранения формат .pickle
    with open('{}_kurs.pickle'.format(i + 1), 'wb') as handle:
        pickle.dump(groups[i], handle, protocol=pickle.HIGHEST_PROTOCOL)
